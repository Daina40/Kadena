import datetime
import json
from django.shortcuts import render
from django.db.models import Sum
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from Production_app.models import ProductionModel
from Production_app.serializers import ProductionSerializer
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from openpyxl import Workbook
import re 
from django.utils.timezone import now
from django.shortcuts import render
from django.db.models import Sum
from rest_framework.views import APIView
import json
from Production_app.models import ProductionModel

class InputListView(APIView):
    def get(self, request, format=None):
        today = now().date()  # Get today's date
        production_list = ProductionModel.objects.filter(date=today)  # Fetch records for today

        # If no data exists for today, ensure an empty list
        if not production_list.exists():
            production_list = []

        # Aggregate totals only if data exists
        if production_list:
            totals = ProductionModel.objects.filter(date=today).aggregate(
                order_qty_sum=Sum('order_qty'),
                daily_prod_sum=Sum('daily_prod'),
                total_prod_sum=Sum('total_prod'),
                daily_insp_sum=Sum('daily_insp'),
                total_insp_sum=Sum('total_insp'),
                daily_packpass_sum=Sum('daily_packpass'),
                total_packpass_sum=Sum('total_packpass'),
                line_balance_sum=Sum('line_balance'),
                insp_balance_sum=Sum('insp_balance'),
                total_balance_sum=Sum('total_balance'),
            )
            totals = {key: (value if value is not None else 0) for key, value in totals.items()}
        else:
            totals = {  # Empty totals for an empty table
                "order_qty_sum": 0,
                "daily_prod_sum": 0,
                "total_prod_sum": 0,
                "daily_insp_sum": 0,
                "total_insp_sum": 0,
                "daily_packpass_sum": 0,
                "total_packpass_sum": 0,
                "line_balance_sum": 0,
                "insp_balance_sum": 0,
                "total_balance_sum": 0,
            }

        # Fetching day, month, year dynamically
        day_name = today.strftime("%A")  # E.g., Wednesday
        month_name = today.strftime("%B")  # E.g., March
        day_number = today.day  # E.g., 5
        year = today.year  # E.g., 2025

        floor_choices = [{"value": key, "text": value} for key, value in ProductionModel.FLOOR_CHOICES]
        line_choices = [{"value": key, "text": value} for key, value in ProductionModel.LINE_CHOICES]

        context = {
            "productions": production_list,
            "day_name": day_name,
            "month_name": month_name,
            "day_number": day_number,
            "year": year,
            "floor_choices": json.dumps(floor_choices),
            "line_choices": json.dumps(line_choices),
            **totals,  # Unpacking total values into context
        }
        return render(request, "daily_input.html", context)



def export_excel(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body).get("table_data", [])

            # Create an in-memory output file for Excel
            output = BytesIO()

            # Create a new workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Production Data"

            # Define column headers
            headers = ["Floor", "Line", "Buyer", "Style", "Item", "PO/Buy", "Order Qty", 
                       "Daily Prod", "Total Prod", "Daily Insp", "Total Insp", "Daily Pack Pass", 
                       "Total Pack Pass", "Line Balance", "Insp Balance", "Total Balance", "Remarks"]
            ws.append(headers)

            # Get the index of the Remarks column
            remarks_col_idx = headers.index("Remarks") + 1  # OpenPyXL uses 1-based index

            # Create mapping dictionaries for Floor and Line
            floor_mapping = dict(ProductionModel.FLOOR_CHOICES)
            line_mapping = dict(ProductionModel.LINE_CHOICES)

            row_num = 2  # Start from second row

            for row in data:
                if row and ("Subtotal" in row[0] or "Final Result" in row[0]):  
                    # Merge cells logic for subtotal rows
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                    merged_cell = ws.cell(row=row_num, column=1, value=row[0])
                    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
                    merged_cell.font = Font(bold=True)

                    for col_idx, value in enumerate(row[1:], start=7):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center")

                else:
                    # Convert Floor & Line values before adding to Excel
                    row[0] = floor_mapping.get(row[0], row[0])  # Convert Floor
                    row[1] = line_mapping.get(row[1], row[1])    # Convert Line

                    # Write normal row data
                    for col_idx, value in enumerate(row, start=1):
                        if col_idx == remarks_col_idx:  # Remarks column only
                            if isinstance(value, list):
                               
                                value = "\n".join(filter(None, value))  # Join remarks with newline
                            else:
                                value = re.sub(r"❌", "", str(value or "")).strip()

                        cell = ws.cell(row=row_num, column=col_idx, value=value)

                        # Enable text wrapping for Remarks column only
                        if col_idx == remarks_col_idx:
                            cell.alignment = Alignment(wrap_text=True)

                row_num += 1

            # Apply border styling to all cells
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # Save the workbook to the BytesIO buffer
            wb.save(output)
            output.seek(0)

            # Create response with correct headers
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="production_data.xlsx"'

            return response

        except Exception as e:
            return HttpResponse(f"Error occurred: {str(e)}", status=500)

    return HttpResponse("Invalid request", status=400)


class ProductionListView(APIView):
    def get(self, request, format=None):
        production_list = ProductionModel.objects.all()

        day_name = request.GET.get('day_name')
        month_name = request.GET.get('month_name')
        day_number = request.GET.get('day_number', 0)
        year = request.GET.get('year', 0)
        
        # If no manual input is provided, use the current date
        if not (day_name and month_name and day_number and year):
            today = datetime.date.today()
            day_name = day_name or today.strftime("%A")
            month_name = month_name or today.strftime("%B")
            day_number = day_number or today.day
            year = year or today.year
        return render(request, 'daily_prod_&_insp.html', {
            'productions': production_list,
            'day_name': day_name,
            'month_name': month_name,
            'day_number': day_number,
            'year': year,
            })


    def post(self, request, format=None):
        # Serialize the incoming data
        serializer = ProductionSerializer(data=request.data)
        if serializer.is_valid():
            # Calculate total_balance before saving
            total_prod = request.data.get('total_prod', 0)
            total_packpass = request.data.get('total_packpass', 0)
            line_balance = request.data.get('line_balance', 0)
            insp_balance = request.data.get('insp_balance', 0)

            # Compute total_balance based on conditions
            if total_prod and total_packpass:
                total_balance = int(total_prod) - int(total_packpass)
            elif line_balance and insp_balance:
                total_balance = int(line_balance) + int(insp_balance)
            else:
                total_balance = 0

            # Add total_balance to the data
            serializer.validated_data['total_balance'] = total_balance

            # Save the instance
            serializer.save()
            return render(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SummaryProductionListView(APIView):
    def get(self, request, format=None):
        productions = ProductionModel.objects.all()

        day_name = request.GET.get('day_name')
        month_name = request.GET.get('month_name')
        day_number = request.GET.get('day_number', 0)
        year = request.GET.get('year', 0)
        
        # If no manual input is provided, use the current date
        if not (day_name and month_name and day_number and year):
            today = datetime.date.today()
            day_name = day_name or today.strftime("%A")
            month_name = month_name or today.strftime("%B")
            day_number = day_number or today.day
            year = year or today.year

        # Calculate subtotals
        subtotals = {
            'order_qty': sum([p.order_qty for p in productions]),
        }

        return render(request, 'summary_of_daily_prod_insp.html', {
            'productions': productions,
            'day_name': day_name,
            'month_name': month_name,
            'day_number': day_number,
            'year': year,
            'subtotals': subtotals,
        })

        # Calculate subtotals for numerical fields
        subtotal_order_qty = sum(p.order_qty for p in productions)
        subtotal_daily_prod = sum(p.daily_prod for p in productions)
        subtotal_total_prod = sum(p.total_prod for p in productions)
        subtotal_daily_insp = sum(p.daily_insp for p in productions)
        subtotal_total_insp = sum(p.total_insp for p in productions)
        subtotal_daily_packpass = sum(p.daily_packpass for p in productions)
        subtotal_total_packpass = sum(p.total_packpass for p in productions)
        subtotal_line_balance = sum(p.line_balance for p in productions)
        subtotal_insp_balance = sum(p.insp_balance for p in productions)
        subtotal_total_balance = sum(p.total_balance for p in productions)
        subtotal_over_prod = sum(p.over_prod for p in productions)
        subtotal_over_pass = sum(p.over_pass for p in productions)
        subtotal_repair_balance = sum(p.repair_balance for p in productions)

        # Render the HTML template
        return render(request, 'summary_of_daily_prod_insp.html', {
            'productions': productions,
            'subtotal_order_qty': subtotal_order_qty,
            'subtotal_daily_prod': subtotal_daily_prod,
            'subtotal_total_prod': subtotal_total_prod,
            'subtotal_daily_insp': subtotal_daily_insp,
            'subtotal_total_insp': subtotal_total_insp,
            'subtotal_daily_packpass': subtotal_daily_packpass,
            'subtotal_total_packpass': subtotal_total_packpass,
            'subtotal_line_balance': subtotal_line_balance,
            'subtotal_insp_balance': subtotal_insp_balance,
            'subtotal_total_balance': subtotal_total_balance,
            'subtotal_over_prod': subtotal_over_prod,
            'subtotal_over_pass': subtotal_over_pass,
            'subtotal_repair_balance': subtotal_repair_balance,
        })

def about(request):
    return render(request, 'about.html')