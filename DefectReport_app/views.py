from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from DefectReport_app.models import DefectStatus
from django.db.models import Sum
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def midline(request):
    queryset = DefectStatus.objects.all().order_by('month')

    # Calculate column-wise totals
    totals = DefectStatus.objects.aggregate(
        total_inspect_qty=Sum('total_inspect_quantity'),
        fabric_fault=Sum('fabric_fault'),
        shading=Sum('shading'),
        trims_missing=Sum('trims_missing'),
        bartack_missing=Sum('bartack_missing'),
        label_stitch_wrong_panel=Sum('label_stitch_wrong_panel'),
        others_1=Sum('others_1'),
        broken=Sum('broken'),
        pleat=Sum('pleat'),
        open_seam=Sum('open_seam'),
        damage_1=Sum('damage_1'),
        needle_mark=Sum('needle_mark'),
        run_off_stitch=Sum('run_off_stitch'),
        high_low=Sum('high_low'),
        bad_tension=Sum('bad_tension'),
        puckering=Sum('puckering'),
        uneven=Sum('uneven'),
        bias=Sum('bias'),
        skip=Sum('skip'),
        twisted=Sum('twisted'),
        fabric_loose=Sum('fabric_loose'),
        fabric_catch=Sum('fabric_catch'),
        zipper_wavy=Sum('zipper_wavy'),
        raw_edge=Sum('raw_edge'),
        others_2=Sum('others_2'),
        out_of_tolerance=Sum('out_of_tolerance'),
        oil=Sum('oil'),
        soilage=Sum('soilage'),
        uncut_thread=Sum('uncut_thread'),
        number_mark=Sum('number_mark'),
        damage_2=Sum('damage_2'),
        dirty_spot=Sum('dirty_spot'),
        color_bleed=Sum('color_bleed'),
        others_defects=Sum('others_defects'),
        total_defect_points=Sum('total_defect_points'),
        total_defect_pcs=Sum('total_defect_pcs'),
    )

    # Calculate DHU and Defective %
    total_inspect_qty = totals['total_inspect_qty'] or 0
    total_defect_points = totals['total_defect_points'] or 0
    total_defect_pcs = totals['total_defect_pcs'] or 0

    dhu_shell_line = (total_defect_points / total_inspect_qty) * 100 if total_inspect_qty else 0
    defective_shell = int((total_defect_pcs / total_inspect_qty) * 100) if total_inspect_qty else 0

    # Add DHU and Defective % to totals
    totals['dhu_shell_line'] = dhu_shell_line
    totals['defective_shell'] = defective_shell

    defect_fields = [
        'fabric_fault', 'shading', 'trims_missing', 'bartack_missing', 'label_stitch_wrong_panel',
        'others_1', 'broken', 'pleat', 'open_seam', 'damage_1', 'needle_mark', 'run_off_stitch',
        'high_low', 'bad_tension', 'puckering', 'uneven', 'bias', 'skip', 'twisted', 'fabric_loose',
        'fabric_catch', 'zipper_wavy', 'raw_edge', 'others_2', 'out_of_tolerance', 'oil', 'soilage',
        'uncut_thread', 'number_mark', 'damage_2', 'dirty_spot', 'color_bleed', 'others_defects'
    ]

    defect_percentages = []
    # Calculate defect percentages
    for field in defect_fields:
        total_field = totals.get(field, 0) or 0
        percentage = (total_field / total_inspect_qty * 100) if total_inspect_qty else 0
        defect_percentages.append({'field': field, 'percentage': percentage})


    return render(request, 'defect_report/midline_defect.html', {
        'queryset': queryset,
        **totals,
        'defect_percentages': defect_percentages
    })

   

def add(request):
    context = {
        'floor_choices': DefectStatus.FLOOR_CHOICES,
        'line_choices': DefectStatus.LINE_CHOICES,
        'total_defect_points': 0,
        'dhu_shell_line': 0,
        'defective_shell': 0,
        
    }
    return render(request, 'defect_report/add.html', context)


@csrf_exempt
def addnow(request):
    if request.method == "POST":
        # Get basic fields
        fields = ['month', 'date', 'floor', 'line', 'qc_name', 'qc_id', 'buyer', 'style', 'buy_po', 'remarks']
        data = {field: request.POST.get(field, '') for field in fields}

        order_qty = int(request.POST.get('order_qty') or 0)

        # Inspected qty and defect pcs
        total_inspect_quantity = int(request.POST.get('total_inspect_quantity') or 0)
        total_defect_pcs = int(request.POST.get('total_defect_pcs') or 0)

        # Defect fields and values
        defect_fields = [
            'fabric_fault', 'shading', 'trims_missing', 'bartack_missing', 'label_stitch_wrong_panel', 'others_1', 'others_2',
            'broken', 'pleat', 'open_seam', 'damage_1', 'damage_2', 'needle_mark', 'run_off_stitch', 'high_low', 'bad_tension',
            'puckering', 'uneven', 'bias', 'skip', 'twisted', 'fabric_loose', 'fabric_catch', 'zipper_wavy',
            'raw_edge', 'out_of_tolerance', 'oil', 'soilage', 'uncut_thread', 'number_mark', 'dirty_spot',
            'color_bleed', 'others_defects'
        ]
        defect_values = {field: int(request.POST.get(field) or 0) for field in defect_fields}
        total_defect_points = sum(defect_values.values())

        # DHU & Defective %
        dhu_shell_line = (total_defect_points / total_inspect_quantity) * 100 if total_inspect_quantity else 0
        defective_shell = int((total_defect_pcs / total_inspect_quantity) * 100) if total_inspect_quantity else 0

        # Save to DB
        DefectStatus.objects.create(
            **data,
            order_qty=order_qty,
            total_inspect_quantity=total_inspect_quantity,
            total_defect_pcs=total_defect_pcs,
            total_defect_points=total_defect_points,
            dhu_shell_line=dhu_shell_line,
            defective_shell=defective_shell,
            **defect_values
        )

        return redirect('midline_defect')

    return redirect('add')


def subtotal(request):
    records = DefectStatus.objects.all()

    context = {
        'total_inspect_qty': records.aggregate(Sum('total_inspect_quantity'))['total_inspect_quantity__sum'] or 0,
        'fabric_fault': records.aggregate(Sum('fabric_fault'))['fabric_fault__sum'] or 0,
        'shading': records.aggregate(Sum('shading'))['shading__sum'] or 0,
        'trims_missing': records.aggregate(Sum('trims_missing'))['trims_missing__sum'] or 0,
        'bartack_missing': records.aggregate(Sum('bartack_missing'))['bartack_missing__sum'] or 0,
        'label_stitch_wrong_panel': records.aggregate(Sum('label_stitch_wrong_panel'))['label_stitch_wrong_panel__sum'] or 0,
        'others_1': records.aggregate(Sum('others_1'))['others_1__sum'] or 0,
        'broken': records.aggregate(Sum('broken'))['broken__sum'] or 0,
        'pleat': records.aggregate(Sum('pleat'))['pleat__sum'] or 0,
        'open_seam': records.aggregate(Sum('open_seam'))['open_seam__sum'] or 0,
        'damage_1': records.aggregate(Sum('damage_1'))['damage_1__sum'] or 0,
        'needle_mark': records.aggregate(Sum('needle_mark'))['needle_mark__sum'] or 0,
        'run_off_stitch': records.aggregate(Sum('run_off_stitch'))['run_off_stitch__sum'] or 0,
        'high_low': records.aggregate(Sum('high_low'))['high_low__sum'] or 0,
        'bad_tension': records.aggregate(Sum('bad_tension'))['bad_tension__sum'] or 0,
        'puckering': records.aggregate(Sum('puckering'))['puckering__sum'] or 0,
        'uneven': records.aggregate(Sum('uneven'))['uneven__sum'] or 0,
        'bias': records.aggregate(Sum('bias'))['bias__sum'] or 0,
        'skip': records.aggregate(Sum('skip'))['skip__sum'] or 0,
        'twisted': records.aggregate(Sum('twisted'))['twisted__sum'] or 0,
        'fabric_loose': records.aggregate(Sum('fabric_loose'))['fabric_loose__sum'] or 0,
        'fabric_catch': records.aggregate(Sum('fabric_catch'))['fabric_catch__sum'] or 0,
        'zipper_wavy': records.aggregate(Sum('zipper_wavy'))['zipper_wavy__sum'] or 0,
        'raw_edge': records.aggregate(Sum('raw_edge'))['raw_edge__sum'] or 0,
        'others_2': records.aggregate(Sum('others_2'))['others_2__sum'] or 0,
        'out_of_tolerance': records.aggregate(Sum('out_of_tolerance'))['out_of_tolerance__sum'] or 0,
        'oil': records.aggregate(Sum('oil'))['oil__sum'] or 0,
        'soilage': records.aggregate(Sum('soilage'))['soilage__sum'] or 0,
        'uncut_thread': records.aggregate(Sum('uncut_thread'))['uncut_thread__sum'] or 0,
        'number_mark': records.aggregate(Sum('number_mark'))['number_mark__sum'] or 0,
        'damage_2': records.aggregate(Sum('damage_2'))['damage_2__sum'] or 0,
        'dirty_spot': records.aggregate(Sum('dirty_spot'))['dirty_spot__sum'] or 0,
        'color_bleed': records.aggregate(Sum('color_bleed'))['color_bleed__sum'] or 0,
        'others_defects': records.aggregate(Sum('others_defects'))['others_defects__sum'] or 0,
        'total_defect_points': records.aggregate(Sum('total_defect_points'))['total_defect_points__sum'] or 0,
        'dhu_shell_line': round((records.aggregate(Sum('total_defect_points'))['total_defect_points__sum'] or 0) /
                                (records.aggregate(Sum('total_inspect_quantity'))['total_inspect_quantity__sum'] or 1) * 100, 2),
        'total_defect_pcs': records.aggregate(Sum('total_defect_pcs'))['total_defect_pcs__sum'] or 0,
        'defective_shell': round((records.aggregate(Sum('total_defect_pcs'))['total_defect_pcs__sum'] or 0) /
                                 (records.aggregate(Sum('total_inspect_quantity'))['total_inspect_quantity__sum'] or 1) * 100, 2),
        'queryset': records 
    }

    return render(request, 'midline_defect.html', context)


@csrf_exempt
def export_midline_defect_excel(request):
    # Queryset and totals
    queryset = DefectStatus.objects.all().order_by('month')
    totals = queryset.aggregate(
        total_inspect_qty=Sum('total_inspect_quantity'),
        fabric_fault=Sum('fabric_fault'),
        shading=Sum('shading'),
        trims_missing=Sum('trims_missing'),
        bartack_missing=Sum('bartack_missing'),
        label_stitch_wrong_panel=Sum('label_stitch_wrong_panel'),
        others_1=Sum('others_1'),
        broken=Sum('broken'),
        pleat=Sum('pleat'),
        open_seam=Sum('open_seam'),
        damage_1=Sum('damage_1'),
        needle_mark=Sum('needle_mark'),
        run_off_stitch=Sum('run_off_stitch'),
        high_low=Sum('high_low'),
        bad_tension=Sum('bad_tension'),
        puckering=Sum('puckering'),
        uneven=Sum('uneven'),
        bias=Sum('bias'),
        skip=Sum('skip'),
        twisted=Sum('twisted'),
        fabric_loose=Sum('fabric_loose'),
        fabric_catch=Sum('fabric_catch'),
        zipper_wavy=Sum('zipper_wavy'),
        raw_edge=Sum('raw_edge'),
        others_2=Sum('others_2'),
        out_of_tolerance=Sum('out_of_tolerance'),
        oil=Sum('oil'),
        soilage=Sum('soilage'),
        uncut_thread=Sum('uncut_thread'),
        number_mark=Sum('number_mark'),
        damage_2=Sum('damage_2'),
        dirty_spot=Sum('dirty_spot'),
        color_bleed=Sum('color_bleed'),
        others_defects=Sum('others_defects'),
        total_defect_points=Sum('total_defect_points'),
        total_defect_pcs=Sum('total_defect_pcs'),
    )

    total_inspect_qty = totals['total_inspect_qty'] or 0
    total_defect_points = totals['total_defect_points'] or 0
    total_defect_pcs = totals['total_defect_pcs'] or 0
    dhu_shell_line = (total_defect_points / total_inspect_qty) * 100 if total_inspect_qty else 0
    defective_shell = (total_defect_pcs / total_inspect_qty) * 100 if total_inspect_qty else 0

    defect_fields = [
        'fabric_fault', 'shading', 'trims_missing', 'bartack_missing', 'label_stitch_wrong_panel',
        'others_1', 'broken', 'pleat', 'open_seam', 'damage_1', 'needle_mark', 'run_off_stitch',
        'high_low', 'bad_tension', 'puckering', 'uneven', 'bias', 'skip', 'twisted', 'fabric_loose',
        'fabric_catch', 'zipper_wavy', 'raw_edge', 'others_2', 'out_of_tolerance', 'oil', 'soilage',
        'uncut_thread', 'number_mark', 'damage_2', 'dirty_spot', 'color_bleed', 'others_defects'
    ]

    defect_percentages = [(field, round((totals.get(field, 0) or 0) / total_inspect_qty * 100, 2)) if total_inspect_qty else (field, 0) for field in defect_fields]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Defect Status"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=48)
    ws['A1'] = "DAILY DEFECT STATUS"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws['A1'].border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )

    # Apply border to all cells in the merged range
    border_style = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )
    for col in range(1, 49):
        cell = ws.cell(row=1, column=col)
        cell.border = border_style

    # Percentage Row
    defect_start_col = 12  # or whatever exact column defect fields begin after your metadata
    for col, (field, percent) in enumerate(defect_percentages, start=defect_start_col):
        ws.cell(row=2, column=col).value = f"{percent}%"
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col).border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        # Ensure subtotals align with defect fields
        ws.cell(row=3, column=col).value = totals.get(field, 0)
        ws.cell(row=3, column=col).font = Font(bold=True)

    # Subtotal Row
    ws.cell(row=3, column=1).value = "Subtotal"
    ws.cell(row=3, column=11).value = total_inspect_qty  # Correctly display Total Inspect Quantity (SHELL)
    ws.cell(row=3, column=11).font = Font(bold=True)

    for col, (field, _) in enumerate(defect_percentages, start=defect_start_col):
        ws.cell(row=3, column=col).value = totals.get(field, 0)
        ws.cell(row=3, column=col).font = Font(bold=True)

        ws.cell(row=3, column=45).value = total_defect_points  # Corrected column index for Total No. Of Defect Points (SHELL)
        ws.cell(row=3, column=45).font = Font(bold=True)
        ws.cell(row=3, column=46).value = round(dhu_shell_line, 2)  # Corrected column index for DHU (SHELL LINE)
        ws.cell(row=3, column=46).font = Font(bold=True)
        ws.cell(row=3, column=46).alignment = Alignment(horizontal="center")
        ws.cell(row=3, column=47).value = totals['total_defect_pcs']  # Corrected to show total_defect_pcs in the subtotal row
        ws.cell(row=3, column=47).font = Font(bold=True)
        ws.cell(row=3, column=48).value = f"{round(defective_shell)}%"  # Ensure subtotal defective % is displayed as a rounded integer with a percentage sign
        ws.cell(row=3, column=48).font = Font(bold=True)
        ws.cell(row=3, column=48).value = round(defective_shell, 2)  # Corrected to show defective_shell in the correct column
        ws.cell(row=3, column=48).font = Font(bold=True)

    # Header Row (Main Header - Row 4)
    main_headers = [
        ("Month", 1), ("Date", 1), ("Floor", 1), ("Line", 1),("Mid Line Qc name", 1), ("Mid Line Qc ID", 1), ("Buyer", 1), ("Style", 1), ("Order Qty", 1), ("Buy/PO", 1),
        ("Total Inspect Quantity (SHELL)", 1), ("A: Fabric", 2), ("B: Trims", 1), ("C: Construction", 3), ("D: Stitching and Sewing", 18),
        ("E: M/ment", 1), ("F: Wash and Trimming", 4), ("G: Laundry wash", 3), ("H: Others", 1), ("Total No. Of Defect Points (SHELL)", 1), ("DHU (SHELL LINE)", 1),
        ("Total Defect Pcs (SHELL)", 1), ("Defective % (SHELL)", 1), ("Remarks", 1)
    ]

    col_cursor = 1
    for title, span in main_headers:
        if span > 1:
            ws.merge_cells(start_row=4, start_column=col_cursor, end_row=4, end_column=col_cursor + span - 1)
        cell = ws.cell(row=4, column=col_cursor)
        cell.value = title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        col_cursor += span

    

    # Subheader Row (Row 5)
    subheaders = [
         "", "", "", "", "", "", "", "", "", "", "", "Fabric Fault", "Shading", "Trims Missing", "Bartack Missing", "Lbl/St Wrong Position", "Others_1", "Broken",
        "Pleat", "Open Seam", "Damage", "Needle Mark", "Run Off Stitch", "High/Low", "Bad Tension", "Puckering",
        "Uneven", "Bias", "Skip", "Twisted", "Fabric Loose", "Fabric Catch", "Zipper Wavy", "Raw Edge", "Others_2",
        "Out Of Tolerance", "Oil", "Soilage", "Uncut Thread", "Number Mark", "Damage", "Dirty Spot",
        "Color Bleed", "Others Defects"
    ]

    for col_index, subheader in enumerate(subheaders, start=1):
        cell = ws.cell(row=5, column=col_index)
        if subheader:  # only add text if there's a subheader
            cell.value = subheader
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                top=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )

    

    # Data rows (from row 6)
    for row_num, obj in enumerate(queryset, start=6):
        ws.cell(row=row_num, column=1).value = obj.month
        ws.cell(row=row_num, column=2).value = obj.date
        ws.cell(row=row_num, column=3).value = obj.floor
        ws.cell(row=row_num, column=4).value = obj.line
        ws.cell(row=row_num, column=5).value = obj.qc_name
        ws.cell(row=row_num, column=6).value = obj.qc_id
        ws.cell(row=row_num, column=7).value = obj.buyer
        ws.cell(row=row_num, column=8).value = obj.style
        ws.cell(row=row_num, column=9).value = obj.order_qty
        ws.cell(row=row_num, column=10).value = obj.buy_po
        ws.cell(row=row_num, column=11).value = obj.total_inspect_quantity

        for col_index, (field, _) in enumerate(defect_percentages, start=12):  # Start from column 12 to align correctly
            ws.cell(row=row_num, column=col_index).value = getattr(obj, field)

        ws.cell(row=row_num, column=45).value = obj.total_defect_points  # Corrected column index for Total No. Of Defect Points (SHELL)
        ws.cell(row=3, column=46).value = round(dhu_shell_line, 2)  # Added DHU (SHELL LINE) column
        ws.cell(row=row_num, column=47).value = obj.total_defect_pcs
        ws.cell(row=row_num, column=48).value = f"{int((obj.total_defect_pcs / obj.total_inspect_quantity * 100))}%" if obj.total_inspect_quantity else "0%"
        ws.cell(row=row_num, column=49).value = obj.remarks

    # Auto-adjust column widths based on the content length
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Daily_Defect_Status.xlsx"'
    wb.save(response)
    return response



# def export_defects_excel(request):
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=defect_report.xlsx'

#     workbook = xlsxwriter.Workbook(response, {'in_memory': True})
#     worksheet = workbook.add_worksheet()

#     # Format
#     header_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

#     # First row: merged headers
#     worksheet.merge_range('A1:B1', 'A: Fabric', header_format)
#     worksheet.merge_range('C1:C1', 'B: Trims', header_format)
#     worksheet.merge_range('D1:F1', 'C: Construction', header_format)
#     worksheet.merge_range('G1:X1', 'D: Stitching and Sewing', header_format)
#     worksheet.merge_range('Y1:Y1', 'E: M/ment', header_format)
#     worksheet.merge_range('Z1:AC1', 'F: Wash and Triming', header_format)
#     worksheet.merge_range('AD1:AF1', 'G: Loundry wash', header_format)
#     worksheet.merge_range('AG1:AG1', 'H: Others', header_format)
#     worksheet.merge_range('AH1:AH1', 'Total No. Of Defect Points (SHELL)', header_format)
#     worksheet.merge_range('AI1:AI1', 'DHU (SHELL LINE)', header_format)
#     worksheet.merge_range('AJ1:AJ1', 'Total Defect Pcs (SHELL)', header_format)
#     worksheet.merge_range('AK1:AK1', 'Defective% (SHEEL)', header_format)
#     worksheet.merge_range('AL1:AL1', 'Remarks', header_format)

#     # Second row: subheaders
#     subheaders = ['Fabric Fault', 'Shading', 'Trims missing', 'Bartack missing', 'Lbl/st wrong position', 'Others', 'Broken', 'Pleat', 'Open seam', 'Damage', 'Needle mark', 'Run off stitch', 'High/low', 'Bad tension', 'Puckering', 'Uneven', 'Bias', 'Skip', 'Twisted', 'Fabric loose', 'Fabric catch', 'Zipper wavy', 'Raw edge', 'Out of tolerance', 'Oil', 'Soilage', 'Uncut thread', 'Number mark', 'Damage 2', 'Dirty spot', 'Color bleed', 'Others defects']
#     worksheet.write_row('A2', subheaders, header_format)

#     # Example data
#     data = [20, 10, 5, 10, 10, 5]
#     worksheet.write_row('A3', data)

#     workbook.close()
#     return response


@csrf_exempt
def update(request, id):
    defect = get_object_or_404(DefectStatus, id=id)

    if request.method == "POST":
        # Update fields
        fields = ['month', 'date', 'floor', 'line', 'qc_name', 'qc_id',
                  'buyer', 'style', 'remarks']
        for field in fields:
            setattr(defect, field, request.POST.get(field, getattr(defect, field)))

        defect.total_inspect_quantity = int(request.POST.get('total_inspect_quantity') or 0)
        defect.total_defect_pcs = int(request.POST.get('total_defect_pcs') or 0)

        defect_fields = [
            'fabric_fault', 'shading', 'trims_missing', 'bartack_missing', 'label_stitch_wrong_panel', 'others_1', 'others_2',
            'broken', 'pleat', 'open_seam', 'damage_1', 'damage_2', 'needle_mark', 'run_off_stitch', 'high_low', 'bad_tension',
            'puckering', 'uneven', 'bias', 'skip', 'twisted', 'fabric_loose', 'fabric_catch', 'zipper_wavy',
            'raw_edge', 'out_of_tolerance', 'oil', 'soilage', 'uncut_thread', 'number_mark', 'dirty_spot',
            'color_bleed', 'others_defects'
        ]

        total_defect_points = 0
        for field in defect_fields:
            value = int(request.POST.get(field) or 0)
            setattr(defect, field, value)
            total_defect_points += value

        defect.total_defect_points = total_defect_points
        defect.dhu_shell_line = round((total_defect_points / defect.total_inspect_quantity) * 100) if defect.total_inspect_quantity else 0
        defect.defective_shell = round((defect.total_defect_pcs / defect.total_inspect_quantity) * 100, 2) if defect.total_inspect_quantity else 0

        defect.save()
        return redirect('midline_defect')

    context = {
        'defect': defect,
        'floor_choices': DefectStatus.FLOOR_CHOICES,
        'line_choices': DefectStatus.LINE_CHOICES,
        'total_defect_points': defect.total_defect_points or 0,
        'dhu_shell_line': defect.dhu_shell_line or 0,
        'defective_shell': defect.defective_shell or 0
    }
    return render(request, 'defect_report/add.html', context)


def delete(request, id):
    defect = get_object_or_404(DefectStatus, id=id)
    defect.delete()
    return redirect('midline_defect')
